# Livro Automatize tarefas maçantes com Python - Al Sweigart
# Capítulo 12

#Importando pacotes
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
import pprint

#Importando arquivo xlsx
wb = openpyxl.load_workbook(r'C:\Users\Dell\Downloads\example.xlsx')

#Obtendo nomes de worksheets
wb.get_sheet_names()

#trabalhando em worksheets específica
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[2])
sheet.title

#active sheet -> A planilha ativa é aquela que estará em evidência quando o workbook for aberto no Excel.
ActiveSheet = wb.active
ActiveSheet.title

#obtendo dados da planilha
sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1'].value

c = sheet['B1']
'Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + str(c.value)
'Cell ' + c.coordinate + ' is ' + c.value

sheet.cell(row=1, column=1)
sheet.cell(row=1, column=1).value

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

#número de linhas da planilha
sheet.max_row

#número de colunas da planilha
sheet.max_column

#convertendo números para letras na coluna
get_column_letter(7)
get_column_letter(77)
get_column_letter(777)

wb = openpyxl.load_workbook(r'C:\Users\Dell\Downloads\example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
get_column_letter(sheet.max_column)

#convertendo letras para números na coluna
column_index_from_string('MM')
column_index_from_string('M')
column_index_from_string('MMM')

#Iteração dentro de intervalo de células
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

#Iterando valores de uma coluna
sheet = wb.active
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)


##############################################################################
#Projeto Censo
wb = openpyxl.load_workbook(r'C:\Users\Dell\Downloads\censuspopdata.xlsx')
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

sheet.max_row
sheet.max_column

countyData = {}
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)
    

print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')

import census2010
census2010.allData['AK']['Anchorage']
anchoragePop = census2010.allData['AK']['Anchorage']['pop']
print('The 2010 population of Anchorage was ' + str(anchoragePop))
##############################################################################

#alterando nome de planilha
wb = openpyxl.load_workbook(r'C:\Users\Dell\Downloads\example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save(r'C:\Users\Dell\Downloads\example.xlsx')

#Criando e removendo planilhas
wb = openpyxl.Workbook()
wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=2, title='Middle Sheet')
wb.create_sheet()
wb.get_sheet_names()

wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
wb.get_sheet_names()

#Escrevendo valores em células
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Hello world!'
sheet['A1'].value